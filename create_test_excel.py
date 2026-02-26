#!/usr/bin/env python3
"""테스트용 Excel 파일 생성"""
from openpyxl import Workbook

# 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "직원명단"

# 헤더
ws['A1'] = "이름"
ws['B1'] = "이메일"
ws['C1'] = "부서"
ws['D1'] = "전화번호"

# 데이터
data = [
    ["홍길동", "hong@example.com", "개발팀", "010-1111-2222"],
    ["김철수", "kim@example.com", "기획팀", "010-3333-4444"],
    ["이영희", "lee@example.com", "디자인팀", "010-5555-6666"],
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 파일 저장
wb.save("/tmp/test_document.xlsx")
print("✅ Excel 파일 생성 완료: /tmp/test_document.xlsx")
