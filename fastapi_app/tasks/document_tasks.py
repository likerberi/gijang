from celery import shared_task
from sqlalchemy.orm import Session
from datetime import datetime
import os
import logging

from .celery_app import celery_app
from ..db.session import SessionLocal
from ..models.document import Document, ExtractedData, Report, MergeProject, MergeFile
from ..schemas.document import DocumentStatus, FileType

logger = logging.getLogger(__name__)


def get_db():
    """데이터베이스 세션"""
    db = SessionLocal()
    try:
        return db
    finally:
        pass


@celery_app.task(bind=True, max_retries=3)
def process_document_task(self, document_id: int):
    """문서 처리 태스크 (Django 로직 재사용)"""
    db = SessionLocal()
    
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            logger.error(f"문서를 찾을 수 없습니다: {document_id}")
            return {"status": "error", "message": "문서를 찾을 수 없습니다"}
        
        # 상태 업데이트
        document.status = DocumentStatus.PROCESSING
        db.commit()
        
        # 파일 유형에 따라 처리
        if document.file_type == FileType.EXCEL:
            result = process_excel(document.file_path)
        elif document.file_type == FileType.IMAGE:
            result = process_image(document.file_path)
        elif document.file_type == FileType.PDF:
            result = process_pdf(document.file_path)
        else:
            raise ValueError(f"지원하지 않는 파일 유형: {document.file_type}")
        
        # 추출된 데이터 저장
        extracted_data = db.query(ExtractedData).filter(
            ExtractedData.document_id == document_id
        ).first()
        
        if extracted_data:
            for key, value in result.items():
                setattr(extracted_data, key, value)
        else:
            extracted_data = ExtractedData(document_id=document_id, **result)
            db.add(extracted_data)
        
        # 문서 상태 업데이트
        document.status = DocumentStatus.COMPLETED
        document.processed_at = datetime.now()
        db.commit()
        
        # 리포트 생성
        generate_report_task.delay(document_id)
        
        logger.info(f"문서 처리 완료: {document.original_filename}")
        return {"status": "success", "document_id": document_id}
        
    except Exception as e:
        logger.error(f"문서 처리 오류: {str(e)}")
        document.status = DocumentStatus.FAILED
        document.error_message = str(e)
        db.commit()
        raise self.retry(exc=e, countdown=60)
    
    finally:
        db.close()


def process_excel(file_path: str) -> dict:
    """엑셀 파일 처리"""
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        headers = data[0] if data else []
        rows = data[1:] if len(data) > 1 else []
        
        structured_data = {
            'headers': headers,
            'rows': rows,
            'sheet_name': sheet.title,
        }
        
        return {
            'extracted_text': str(data),
            'structured_data': structured_data,
            'total_rows': len(rows),
            'meta_info': {  # metadata -> meta_info로 변경
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
            }
        }
    except Exception as e:
        logger.error(f"엑셀 처리 오류: {str(e)}")
        raise


def process_image(file_path: str) -> dict:
    """이미지 파일 처리"""
    try:
        from PIL import Image
        
        img = Image.open(file_path)
        
        metadata_dict = {
            'format': img.format,
            'mode': img.mode,
            'size': img.size,
            'width': img.width,
            'height': img.height,
        }
        
        return {
            'extracted_text': '',
            'structured_data': {},
            'meta_info': metadata_dict,  # metadata -> meta_info로 변경
        }
    except Exception as e:
        logger.error(f"이미지 처리 오류: {str(e)}")
        raise


def process_pdf(file_path: str) -> dict:
    """PDF 파일 처리"""
    try:
        from PyPDF2 import PdfReader
        
        reader = PdfReader(file_path)
        
        text = ''
        for page in reader.pages:
            text += page.extract_text() + '\n'
        
        return {
            'extracted_text': text,
            'structured_data': {},
            'total_pages': len(reader.pages),
            'meta_info': {  # metadata -> meta_info로 변경
                'page_count': len(reader.pages),
            }
        }
    except Exception as e:
        logger.error(f"PDF 처리 오류: {str(e)}")
        raise


@celery_app.task
def generate_report_task(document_id: int):
    """리포트 생성 태스크"""
    db = SessionLocal()
    
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            return
        
        extracted_data = db.query(ExtractedData).filter(
            ExtractedData.document_id == document_id
        ).first()
        
        if not extracted_data:
            return
        
        # 리포트 생성
        title = f"{document.original_filename} 분석 리포트"
        summary = generate_summary(extracted_data)
        
        content = {
            'file_info': {
                'filename': document.original_filename,
                'file_type': document.file_type.value,
                'file_size': document.file_size,
                'processed_at': document.processed_at.isoformat() if document.processed_at else None,
            },
            'extracted_data': {
                'total_pages': extracted_data.total_pages,
                'total_rows': extracted_data.total_rows,
                'text_length': len(extracted_data.extracted_text),
            },
            'structured_data': extracted_data.structured_data,
        }
        
        report = Report(
            document_id=document_id,
            title=title,
            summary=summary,
            content=content,
            generated_by=document.user_id,
        )
        
        db.add(report)
        db.commit()
        
        logger.info(f"리포트 생성 완료: {document.original_filename}")
        
    except Exception as e:
        logger.error(f"리포트 생성 오류: {str(e)}")
        raise
    
    finally:
        db.close()


def generate_summary(extracted_data: ExtractedData) -> str:
    """요약 생성"""
    summary_parts = []
    
    if extracted_data.total_pages > 0:
        summary_parts.append(f"총 {extracted_data.total_pages}페이지")
    
    if extracted_data.total_rows > 0:
        summary_parts.append(f"총 {extracted_data.total_rows}행의 데이터")
    
    text_length = len(extracted_data.extracted_text)
    if text_length > 0:
        summary_parts.append(f"{text_length:,}자의 텍스트 추출됨")
    
    return ", ".join(summary_parts) if summary_parts else "데이터 추출 완료"


# ========================
# 파일 병합 태스크
# ========================

@celery_app.task(bind=True, max_retries=2)
def analyze_merge_files_task(self, project_id: int):
    """병합 프로젝트 파일 분석 태스크"""
    db = SessionLocal()
    
    try:
        project = db.query(MergeProject).filter(MergeProject.id == project_id).first()
        if not project:
            logger.error(f"병합 프로젝트를 찾을 수 없음: {project_id}")
            return {"status": "error"}
        
        project.status = "analyzing"
        db.commit()
        
        # documents 앱의 유틸리티 재사용
        import sys
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        from documents.utils.merge_service import MergeService
        
        service = MergeService()
        merge_files = db.query(MergeFile).filter(MergeFile.project_id == project_id).all()
        file_paths = [mf.file_path for mf in merge_files]
        
        analysis = service.analyze_files(file_paths)
        
        # 개별 파일 결과 저장
        for file_info in analysis.get('files', []):
            for mf in merge_files:
                if mf.original_filename == file_info.get('filename'):
                    if 'error' not in file_info:
                        mf.detected_headers = file_info.get('headers', [])
                        mf.header_row_index = file_info.get('header_row_index', 0)
                        mf.total_rows = file_info.get('total_rows', 0)
                        mf.column_types = file_info.get('column_types', {})
                        mf.sample_data = file_info.get('sample_data', [])
                        mf.is_analyzed = 1
                    else:
                        mf.error_message = file_info['error']
                    break
        
        project.analysis_result = {
            'suggested_mappings': analysis.get('suggested_mappings', {}),
            'all_headers': analysis.get('all_headers', []),
            'analyzed_at': datetime.now().isoformat(),
        }
        project.status = "ready"
        db.commit()
        
        logger.info(f"병합 분석 완료: {project.name}")
        return {"status": "success", "project_id": project_id}
        
    except Exception as e:
        logger.error(f"병합 분석 오류: {str(e)}")
        try:
            project.status = "failed"
            project.error_message = str(e)
            db.commit()
        except Exception:
            pass
        raise self.retry(exc=e, countdown=30)
    finally:
        db.close()


@celery_app.task(bind=True, max_retries=2)
def execute_merge_task(self, project_id: int):
    """병합 실행 태스크"""
    db = SessionLocal()
    
    try:
        project = db.query(MergeProject).filter(MergeProject.id == project_id).first()
        if not project:
            return {"status": "error"}
        
        project.status = "merging"
        project.error_message = None
        db.commit()
        
        import sys
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        from documents.utils.merge_service import MergeService
        from documents.utils.normalizers import DateNormalizer
        
        service = MergeService()
        
        if project.date_output_format:
            service.date_normalizer = DateNormalizer(output_format=project.date_output_format)
        
        if project.column_mapping:
            for original, standard in project.column_mapping.items():
                service.column_mapper.add_mapping(standard, [original])
        
        merge_files = db.query(MergeFile).filter(MergeFile.project_id == project_id).all()
        file_paths = [mf.file_path for mf in merge_files]
        
        output_dir = os.path.join(
            str(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            'media', 'merged', datetime.now().strftime('%Y/%m/%d')
        )
        os.makedirs(output_dir, exist_ok=True)
        
        output_filename = f'merged_{project_id}_{datetime.now().strftime("%H%M%S")}.xlsx'
        output_path = os.path.join(output_dir, output_filename)
        
        result = service.merge_files(
            file_paths=file_paths,
            column_mapping=project.column_mapping or {},
            date_columns=project.date_columns or [],
            number_columns=project.number_columns or [],
            output_path=output_path,
            add_source_column=True,
        )
        
        if result['success']:
            project.merged_file_path = output_path
            project.merge_log = result
            project.status = "completed"
            project.completed_at = datetime.now()
            
            for log_entry in result.get('merge_log', []):
                for mf in merge_files:
                    if mf.original_filename == log_entry.get('file'):
                        mf.is_processed = 1 if log_entry.get('status') == 'success' else 0
                        if log_entry.get('error'):
                            mf.error_message = log_entry['error']
                        break
        else:
            project.status = "failed"
            project.error_message = result.get('error', '병합 실패')
            project.merge_log = result
        
        db.commit()
        logger.info(f"병합 실행 완료: {project.name}")
        return {"status": "success", "project_id": project_id}
        
    except Exception as e:
        logger.error(f"병합 실행 오류: {str(e)}")
        try:
            project.status = "failed"
            project.error_message = str(e)
            db.commit()
        except Exception:
            pass
        raise self.retry(exc=e, countdown=30)
    finally:
        db.close()
