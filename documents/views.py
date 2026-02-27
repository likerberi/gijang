from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.db.models import F
from .models import Document, ExtractedData, Report, MergeProject, MergeFile, ColumnMappingTemplate, Vendor, ClassificationRule
from .serializers import (
    DocumentSerializer, DocumentUploadSerializer,
    ExtractedDataSerializer, ReportSerializer,
    MergeProjectSerializer, MergeProjectCreateSerializer,
    MergeProjectUpdateMappingSerializer, MergeFileSerializer,
    MergeFileUploadSerializer, ColumnMappingTemplateSerializer,
    VendorSerializer,
)
from .tasks import process_document, analyze_merge_files, execute_merge
import math
import logging
import time
from datetime import datetime, date
from collections import defaultdict

logger = logging.getLogger(__name__)

# Celery worker 상태 캐시 (60초)
_worker_cache = {'available': None, 'checked_at': 0}


def _is_celery_worker_available():
    """Celery worker가 실행 중인지 확인 (60초 캐싱)"""
    global _worker_cache
    now = time.time()
    if _worker_cache['available'] is not None and (now - _worker_cache['checked_at']) < 60:
        return _worker_cache['available']
    try:
        from config.celery import app
        result = app.control.ping(timeout=1.0)
        available = bool(result)
    except Exception:
        available = False
    _worker_cache['available'] = available
    _worker_cache['checked_at'] = now
    if not available:
        logger.info("Celery worker 미감지 → 동기 모드로 전환")
    return available


def _dispatch_task(task, *args, **kwargs):
    """Celery 태스크 디스패치 — worker 없으면 동기 실행"""
    if _is_celery_worker_available():
        return task.delay(*args, **kwargs)
    logger.info(f"동기 실행: {task.name}({args})")
    try:
        return task.apply(args=args, kwargs=kwargs)
    except Exception as e:
        logger.error(f"동기 실행 실패: {e}", exc_info=True)
        raise


# ========================
# 기장용 헬퍼 함수
# ========================

# 기장에서 흔히 쓰는 계정과목 매핑
ACCOUNT_CATEGORY_KEYWORDS = {
    '매출': ['매출', '판매', '수입', '용역', '매출 입금'],
    '이자수익': ['이자수입', '예금이자', '이자수익'],
    '잡수입': ['잡수입', '잡이익', '기타수입', '임대수입'],
    '급여': ['급여', '월급', '상여', '보너스', '인건비', '퇴직금'],
    '복리후생비': ['복리후생', '식대', '직원식대', '체력단련', '경조사', '건강검진'],
    '소모품비': ['소모품', '사무용품', '비품'],
    '접대비': ['접대', '식비접대', '거래처 식비', '거래처 선물', '접대비'],
    '여비교통비': ['교통', '출장', '택시', '주유', '기차', 'KTX', '버스', '항공', '유류', '승차권'],
    '통신비': ['통신', '전화', '인터넷', '핸드폰', '휴대폰'],
    '수도광열비': ['수도', '전기', '가스', '광열', '전기요금', '가스요금'],
    '임차료': ['임대료', '월세', '임차', '사무실 임대'],
    '보험료': ['보험', '산재', '고용보험', '건강보험', '국민연금'],
    '세금과공과': ['세금', '부가세', '법인세', '국세', '지방세', '인지세', '등록세', '중간예납', '예정신고'],
    '광고선전비': ['광고', '홍보', '마케팅', '판촉', '광고비'],
    '수수료': ['수수료', '카드수수료', '이체수수료', '계좌이체'],
    '이자비용': ['이자', '대출이자'],
    '감가상각비': ['감가상각'],
    '외주비': ['외주', '용역비', '하도급', '외주비'],
    '수선비': ['수선', '수리', '유지보수'],
}

# 입금/출금/잔액으로 판단 가능한 열 이름 패턴
FINANCIAL_COLUMN_PATTERNS = {
    'income': ['입금', '입금액', '수입', '매출', '수입금액', '대변', 'credit', 'income', '받은금액'],
    'expense': ['출금', '출금액', '지출', '지급', '차변', 'debit', 'expense', '보낸금액', '사용금액', '결제금액', '이용금액'],
    'balance': ['잔액', '잔고', '누적잔액', 'balance', '계좌잔액'],
    'date': ['날짜', '거래일', '일자', '거래일자', '사용일', '이용일', '승인일', '결제일', 'date', '거래일시'],
    'description': ['적요', '내용', '거래내용', '비고', '메모', '상세', '거래처', '이용내역', '사용처', '가맹점', '거래적요'],
}


def detect_financial_columns(headers):
    """헤더에서 입금/출금/잔액/날짜/적요 열을 자동 감지"""
    result = {}
    if not headers:
        return result
    
    normalized_headers = [str(h).strip().lower() for h in headers]
    
    for col_type, patterns in FINANCIAL_COLUMN_PATTERNS.items():
        for i, header in enumerate(normalized_headers):
            for pattern in patterns:
                if pattern.lower() in header:
                    result[col_type] = i
                    break
            if col_type in result:
                break
    
    return result


def classify_transaction(description):
    """적요 내용으로 계정과목 자동 분류"""
    if not description:
        return '미분류'
    desc = str(description).strip()
    for category, keywords in ACCOUNT_CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in desc:
                return category
    return '미분류'


def classify_transaction_with_rules(description, user=None):
    """사용자 학습 규칙을 우선 적용하는 분류 함수
    
    우선순위: 사용자 exact 매칭 → 사용자 contains 매칭 → 키워드 기본 분류
    """
    if not description:
        return '미분류'
    desc = str(description).strip()
    
    # 1) 사용자 학습 규칙 적용 (우선순위 순)
    if user:
        rules = ClassificationRule.objects.filter(
            user=user, is_active=True
        ).order_by('priority', '-hit_count')
        
        for rule in rules:
            if rule.match_type == 'exact' and rule.pattern == desc:
                ClassificationRule.objects.filter(pk=rule.pk).update(
                    hit_count=F('hit_count') + 1
                )
                return rule.category
            elif rule.match_type == 'contains' and rule.pattern in desc:
                ClassificationRule.objects.filter(pk=rule.pk).update(
                    hit_count=F('hit_count') + 1
                )
                return rule.category
            elif rule.match_type == 'vendor' and rule.pattern in desc:
                ClassificationRule.objects.filter(pk=rule.pk).update(
                    hit_count=F('hit_count') + 1
                )
                return rule.category
    
    # 2) 기본 키워드 분류
    return classify_transaction(desc)


def compute_financial_summary(structured_data):
    """구조화된 데이터에서 재무 요약 계산"""
    headers = structured_data.get('headers', [])
    rows = structured_data.get('rows', [])
    
    if not headers or not rows:
        return None
    
    fin_cols = detect_financial_columns(headers)
    if not fin_cols:
        return None
    
    summary = {
        'detected_columns': {k: headers[v] if v < len(headers) else None for k, v in fin_cols.items()},
        'total_income': 0,
        'total_expense': 0,
        'net': 0,
        'transaction_count': len(rows),
        'category_breakdown': {},
    }
    
    desc_idx = fin_cols.get('description')
    income_idx = fin_cols.get('income')
    expense_idx = fin_cols.get('expense')
    
    for row in rows:
        # 입금 합계
        if income_idx is not None and income_idx < len(row):
            val = row[income_idx]
            try:
                amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                if not math.isnan(amount):
                    summary['total_income'] += amount
            except (ValueError, TypeError):
                pass
        
        # 출금 합계
        if expense_idx is not None and expense_idx < len(row):
            val = row[expense_idx]
            try:
                amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                if not math.isnan(amount):
                    summary['total_expense'] += amount
            except (ValueError, TypeError):
                pass
        
        # 계정과목 분류
        if desc_idx is not None and desc_idx < len(row):
            category = classify_transaction(row[desc_idx])
            if category not in summary['category_breakdown']:
                summary['category_breakdown'][category] = {'count': 0, 'income': 0, 'expense': 0}
            summary['category_breakdown'][category]['count'] += 1
            
            if income_idx is not None and income_idx < len(row):
                try:
                    val = row[income_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if not math.isnan(amount):
                        summary['category_breakdown'][category]['income'] += amount
                except (ValueError, TypeError):
                    pass
            if expense_idx is not None and expense_idx < len(row):
                try:
                    val = row[expense_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if not math.isnan(amount):
                        summary['category_breakdown'][category]['expense'] += amount
                except (ValueError, TypeError):
                    pass
    
    summary['net'] = summary['total_income'] - summary['total_expense']
    return summary


def _parse_date(date_str):
    """다양한 날짜 형식을 파싱"""
    if not date_str:
        return None
    date_str = str(date_str).strip()
    formats = [
        '%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d',
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
        '%Y.%m.%d %H:%M:%S', '%Y.%m.%d %H:%M',
        '%m/%d/%Y', '%d-%m-%Y',
        '%Y%m%d',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    # 숫자만 있는 경우 (Excel serial date)
    try:
        serial = float(date_str)
        if 30000 < serial < 60000:
            from datetime import timedelta
            return datetime(1899, 12, 30) + timedelta(days=int(serial))
    except:
        pass
    return None


def _extract_vendor_name(description):
    """적요에서 거래처명 추출"""
    if not description:
        return ''
    desc = str(description).strip()
    
    # 입금/출금 접두어 등 제거
    prefixes_to_remove = [
        '매출 입금 - ', '매출 입금-', '입금 - ', '입금-',
        '출금 - ', '출금-', '이체 - ', '이체-',
        '카드결제 - ', '카드결제-', '체크카드 ', '신용카드 ',
        '자동이체 ', '급여이체 ', 'CMS출금 ', 'CMS ',
    ]
    for prefix in prefixes_to_remove:
        if desc.startswith(prefix):
            desc = desc[len(prefix):]
            break
    
    # 괄호 안 내용 제거
    import re
    desc = re.sub(r'\(.*?\)', '', desc).strip()
    desc = re.sub(r'\[.*?\]', '', desc).strip()
    
    # 날짜/시간 패턴 제거
    desc = re.sub(r'\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', '', desc).strip()
    desc = re.sub(r'\d{1,2}:\d{2}(:\d{2})?', '', desc).strip()
    
    # 과도한 공백 정리
    desc = re.sub(r'\s+', ' ', desc).strip()
    
    return desc if len(desc) >= 2 else ''


# ========================
# 세금 달력 데이터
# ========================

TAX_CALENDAR = [
    {'month': 1, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 1, 'day': 25, 'title': '부가가치세 확정신고', 'desc': '7~12월분 (2기) 확정신고·납부', 'type': 'quarterly'},
    {'month': 1, 'day': 31, 'title': '지급명세서 제출', 'desc': '전년도 근로·사업소득 지급명세서', 'type': 'annual'},
    {'month': 2, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 3, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 3, 'day': 31, 'title': '법인세 신고·납부', 'desc': '12월 결산법인 법인세 신고·납부', 'type': 'annual'},
    {'month': 4, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 4, 'day': 25, 'title': '부가가치세 예정신고', 'desc': '1~3월분 (1기) 예정신고·납부', 'type': 'quarterly'},
    {'month': 5, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 5, 'day': 31, 'title': '종합소득세 신고·납부', 'desc': '전년도 종합소득세 확정신고·납부', 'type': 'annual'},
    {'month': 6, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 7, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 7, 'day': 25, 'title': '부가가치세 확정신고', 'desc': '1~6월분 (1기) 확정신고·납부', 'type': 'quarterly'},
    {'month': 8, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 8, 'day': 31, 'title': '법인세 중간예납', 'desc': '12월 결산법인 중간예납', 'type': 'annual'},
    {'month': 9, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 10, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 10, 'day': 25, 'title': '부가가치세 예정신고', 'desc': '7~9월분 (2기) 예정신고·납부', 'type': 'quarterly'},
    {'month': 11, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
    {'month': 11, 'day': 30, 'title': '종합소득세 중간예납', 'desc': '종합소득세 중간예납 납부', 'type': 'annual'},
    {'month': 12, 'day': 10, 'title': '원천세 신고·납부', 'desc': '전월분 원천징수세액 신고·납부', 'type': 'monthly'},
]


class DocumentViewSet(viewsets.ModelViewSet):
    """문서 뷰셋"""
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['file_type', 'status']
    search_fields = ['original_filename']
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return Document.objects.filter(user=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return DocumentUploadSerializer
        return DocumentSerializer
    
    def perform_create(self, serializer):
        document = serializer.save()
        _dispatch_task(process_document, document.id)
    
    @action(detail=True, methods=['post'])
    def reprocess(self, request, pk=None):
        """문서 재처리"""
        document = self.get_object()
        
        if document.status == 'processing':
            return Response(
                {'error': '이미 처리 중인 문서입니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        document.status = 'pending'
        document.error_message = ''
        document.save()
        
        _dispatch_task(process_document, document.id)
        
        return Response({'message': '문서 재처리가 시작되었습니다.'})
    
    @action(detail=True, methods=['get'])
    def extracted_data(self, request, pk=None):
        """추출된 데이터 조회"""
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
            serializer = ExtractedDataSerializer(extracted)
            return Response(serializer.data)
        except ExtractedData.DoesNotExist:
            return Response(
                {'error': '추출된 데이터가 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['get'])
    def reports(self, request, pk=None):
        """문서의 리포트 목록"""
        document = self.get_object()
        reports = document.reports.all()
        serializer = ReportSerializer(reports, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def data(self, request, pk=None):
        """전체 데이터 페이지네이션 조회 (기장용)
        
        쿼리 파라미터:
        - page: 페이지 번호 (기본 1)
        - page_size: 페이지당 행 수 (기본 50, 최대 500)
        - search: 적요/내용 검색어
        - sort_col: 정렬 열 인덱스
        - sort_dir: 정렬 방향 (asc/desc)
        - category: 계정과목 필터
        """
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response(
                {'error': '추출된 데이터가 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        
        if not headers:
            return Response({'headers': [], 'rows': [], 'total': 0, 'financial_summary': None})
        
        # 재무 열 감지
        fin_cols = detect_financial_columns(headers)
        desc_idx = fin_cols.get('description')
        
        # 검색 필터
        search = request.query_params.get('search', '').strip()
        if search:
            filtered = []
            for row in rows:
                match = any(search.lower() in str(cell).lower() for cell in row)
                if match:
                    filtered.append(row)
            rows = filtered
        
        # 계정과목 필터
        category_filter = request.query_params.get('category', '').strip()
        if category_filter and desc_idx is not None:
            filtered = []
            for row in rows:
                if desc_idx < len(row):
                    cat = classify_transaction(row[desc_idx])
                    if cat == category_filter:
                        filtered.append(row)
            rows = filtered
        
        total = len(rows)
        
        # 정렬
        sort_col = request.query_params.get('sort_col', '')
        sort_dir = request.query_params.get('sort_dir', 'asc')
        if sort_col:
            try:
                col_idx = int(sort_col)
                if 0 <= col_idx < len(headers):
                    def sort_key(r):
                        val = r[col_idx] if col_idx < len(r) else None
                        if val is None or val == '' or val == '-':
                            # None/빈칸은 항상 맨 뒤로
                            return (1, 0, '')
                        s = str(val).replace(',', '').replace('원', '').strip()
                        try:
                            return (0, float(s), '')
                        except (ValueError, TypeError):
                            return (0, float('inf'), s.lower())
                    rows = sorted(rows, key=sort_key, reverse=(sort_dir == 'desc'))
            except (ValueError, IndexError):
                pass
        
        # 페이지네이션
        page = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 50)), 500)
        start = (page - 1) * page_size
        end = start + page_size
        
        # 각 행에 계정과목 분류 추가 (사용자 학습 규칙 우선)
        page_rows = []
        user_classifications = (extracted.metadata or {}).get('user_classifications', {})
        global_start = start  # 전체 데이터에서의 시작 인덱스
        for i, row in enumerate(rows[start:end]):
            row_global_idx = start + i
            # 1) 사용자가 수동 지정한 분류 우선
            if str(row_global_idx) in user_classifications:
                cat = user_classifications[str(row_global_idx)]
            elif desc_idx is not None and desc_idx < len(row):
                cat = classify_transaction_with_rules(row[desc_idx], user=request.user)
            else:
                cat = '미분류'
            page_rows.append({
                'cells': row,
                'category': cat,
                'user_classified': str(row_global_idx) in user_classifications,
            })
        
        # 재무 요약
        financial_summary = compute_financial_summary(sd)
        
        return Response({
            'headers': headers,
            'rows': page_rows,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': math.ceil(total / page_size) if total > 0 else 0,
            'financial_columns': {k: headers[v] if v < len(headers) else None for k, v in fin_cols.items()},
            'financial_summary': financial_summary,
            'sheet_name': sd.get('sheet_name', ''),
        })
    
    @action(detail=True, methods=['get'])
    def download_data(self, request, pk=None):
        """추출 데이터를 엑셀로 다운로드 (기장용)
        
        계정과목 분류 열을 추가하여 가공된 엑셀 파일 생성
        """
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from io import BytesIO
        
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response(
                {'error': '추출된 데이터가 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        
        if not headers:
            return Response(
                {'error': '구조화된 데이터가 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 재무 열 감지
        fin_cols = detect_financial_columns(headers)
        desc_idx = fin_cols.get('description')
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        
        # === 시트 1: 거래내역 ===
        ws = wb.active
        ws.title = '거래내역'
        
        # 헤더 (+ 계정과목 열 추가)
        export_headers = list(headers) + ['계정과목']
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_idx, h in enumerate(export_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(h) if h else '')
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 데이터
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                # 숫자 열은 천단위 구분
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
            
            # 계정과목 열 추가
            cat = '미분류'
            if desc_idx is not None and desc_idx < len(row):
                cat = classify_transaction(row[desc_idx])
            ws.cell(row=row_idx, column=len(export_headers), value=cat)
        
        # 열 너비 자동 조정
        for col_idx in range(1, len(export_headers) + 1):
            max_len = len(str(export_headers[col_idx - 1] or ''))
            for row_idx in range(2, min(len(rows) + 2, 100)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)
        
        # === 시트 2: 요약 ===
        financial_summary = compute_financial_summary(sd)
        if financial_summary:
            ws2 = wb.create_sheet('재무요약')
            
            # 제목
            ws2.cell(row=1, column=1, value=f'{document.original_filename} 재무 요약').font = Font(bold=True, size=14)
            ws2.merge_cells('A1:D1')
            
            # 기본 통계
            ws2.cell(row=3, column=1, value='구분').font = Font(bold=True)
            ws2.cell(row=3, column=2, value='금액').font = Font(bold=True)
            
            ws2.cell(row=4, column=1, value='총 입금')
            ws2.cell(row=4, column=2, value=financial_summary['total_income']).number_format = '#,##0'
            ws2.cell(row=5, column=1, value='총 출금')
            ws2.cell(row=5, column=2, value=financial_summary['total_expense']).number_format = '#,##0'
            ws2.cell(row=6, column=1, value='순이익').font = Font(bold=True)
            ws2.cell(row=6, column=2, value=financial_summary['net']).number_format = '#,##0'
            ws2.cell(row=6, column=2).font = Font(bold=True)
            ws2.cell(row=7, column=1, value='거래 건수')
            ws2.cell(row=7, column=2, value=financial_summary['transaction_count'])
            
            # 계정과목별 요약
            breakdown = financial_summary.get('category_breakdown', {})
            if breakdown:
                ws2.cell(row=9, column=1, value='계정과목별 요약').font = Font(bold=True, size=12)
                
                cat_headers = ['계정과목', '건수', '입금', '출금', '순액']
                for ci, ch in enumerate(cat_headers, 1):
                    ws2.cell(row=10, column=ci, value=ch).font = Font(bold=True)
                    ws2.cell(row=10, column=ci).fill = PatternFill(start_color='F1F5F9', fill_type='solid')
                
                for ri, (cat, info) in enumerate(sorted(breakdown.items()), 11):
                    ws2.cell(row=ri, column=1, value=cat)
                    ws2.cell(row=ri, column=2, value=info['count'])
                    ws2.cell(row=ri, column=3, value=info['income']).number_format = '#,##0'
                    ws2.cell(row=ri, column=4, value=info['expense']).number_format = '#,##0'
                    ws2.cell(row=ri, column=5, value=info['income'] - info['expense']).number_format = '#,##0'
            
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 15
            ws2.column_dimensions['D'].width = 15
            ws2.column_dimensions['E'].width = 15
        
        # 응답
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        filename = f'{document.original_filename.rsplit(".", 1)[0]}_기장정리.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """재무 요약 API (대시보드용)"""
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response(
                {'error': '추출된 데이터가 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        sd = extracted.structured_data
        financial_summary = compute_financial_summary(sd)
        
        return Response({
            'document_id': document.id,
            'filename': document.original_filename,
            'total_rows': extracted.total_rows,
            'financial_summary': financial_summary,
        })
    
    @action(detail=True, methods=['post'])
    def classify(self, request, pk=None):
        """거래 분류 수동 업데이트 + 학습
        
        body: { "classifications": { "row_index": "계정과목", ... } }
        사용자가 수동으로 지정한 계정과목을 저장하고,
        적요 패턴을 ClassificationRule로 학습시킴
        """
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response(
                {'error': '추출된 데이터가 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        classifications = request.data.get('classifications', {})
        if not classifications:
            return Response(
                {'error': 'classifications 데이터가 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # metadata에 사용자 분류 저장
        meta = extracted.metadata or {}
        user_classifications = meta.get('user_classifications', {})
        user_classifications.update(classifications)
        meta['user_classifications'] = user_classifications
        extracted.metadata = meta
        extracted.save()
        
        # ★ 학습: 적요 → 계정과목 매핑을 ClassificationRule에 저장
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        fin_cols = detect_financial_columns(headers)
        desc_idx = fin_cols.get('description')
        
        rules_created = 0
        rules_updated = 0
        
        if desc_idx is not None:
            for row_index_str, category in classifications.items():
                try:
                    row_idx = int(row_index_str)
                    if 0 <= row_idx < len(rows) and desc_idx < len(rows[row_idx]):
                        desc = str(rows[row_idx][desc_idx]).strip()
                        if desc and len(desc) >= 2:
                            rule, created = ClassificationRule.objects.update_or_create(
                                user=request.user,
                                pattern=desc,
                                match_type='exact',
                                defaults={
                                    'category': category,
                                    'source': 'user',
                                    'priority': 10,
                                }
                            )
                            if created:
                                rules_created += 1
                            else:
                                rules_updated += 1
                                rule.hit_count += 1
                                rule.save()
                except (ValueError, IndexError):
                    pass
        
        return Response({
            'message': f'{len(classifications)}건의 분류가 저장되었습니다.',
            'total_classified': len(user_classifications),
            'rules_learned': rules_created + rules_updated,
        })

    # ========================
    # 부가세 신고서
    # ========================

    @action(detail=True, methods=['get'])
    def vat_report(self, request, pk=None):
        """부가세 신고 데이터 생성
        
        쿼리 파라미터:
        - quarter: 분기 (1,2,3,4) - 미지정시 전체
        - year: 연도 - 미지정시 현재년
        """
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response({'error': '추출된 데이터가 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        
        if not headers or not rows:
            return Response({'error': '데이터가 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        fin_cols = detect_financial_columns(headers)
        income_idx = fin_cols.get('income')
        expense_idx = fin_cols.get('expense')
        desc_idx = fin_cols.get('description')
        date_idx = fin_cols.get('date')
        
        # 분기 필터
        quarter = request.query_params.get('quarter')
        year = request.query_params.get('year', str(date.today().year))
        
        filtered_rows = rows
        if quarter and date_idx is not None:
            q = int(quarter)
            q_months = {1: [1,2,3], 2: [4,5,6], 3: [7,8,9], 4: [10,11,12]}
            target_months = q_months.get(q, [])
            filtered = []
            for row in rows:
                if date_idx < len(row) and row[date_idx]:
                    try:
                        d = _parse_date(str(row[date_idx]))
                        if d and d.month in target_months and str(d.year) == year:
                            filtered.append(row)
                    except:
                        filtered.append(row)
                else:
                    filtered.append(row)
            filtered_rows = filtered
        
        # 매출/매입 분류
        sales_items = []  # 매출 (수입)
        purchase_items = []  # 매입 (지출)
        
        total_sales = 0
        total_purchases = 0
        
        for row in filtered_rows:
            desc = str(row[desc_idx]) if desc_idx is not None and desc_idx < len(row) else ''
            category = classify_transaction(desc)
            
            inc_amount = 0
            exp_amount = 0
            
            if income_idx is not None and income_idx < len(row):
                try:
                    val = row[income_idx]
                    inc_amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if math.isnan(inc_amount):
                        inc_amount = 0
                except:
                    inc_amount = 0
            
            if expense_idx is not None and expense_idx < len(row):
                try:
                    val = row[expense_idx]
                    exp_amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if math.isnan(exp_amount):
                        exp_amount = 0
                except:
                    exp_amount = 0
            
            date_str = str(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
            
            if inc_amount > 0:
                total_sales += inc_amount
                sales_items.append({
                    'date': date_str,
                    'description': desc,
                    'category': category,
                    'amount': inc_amount,
                    'vat': round(inc_amount / 11, 0),  # 부가세 포함가 기준
                    'supply': round(inc_amount * 10 / 11, 0),
                })
            
            if exp_amount > 0:
                total_purchases += exp_amount
                purchase_items.append({
                    'date': date_str,
                    'description': desc,
                    'category': category,
                    'amount': exp_amount,
                    'vat': round(exp_amount / 11, 0),
                    'supply': round(exp_amount * 10 / 11, 0),
                })
        
        # 세액 계산
        sales_vat = round(total_sales / 11, 0)
        purchase_vat = round(total_purchases / 11, 0)
        payable_vat = sales_vat - purchase_vat
        
        # 카테고리별 매입 합계
        purchase_by_category = defaultdict(lambda: {'count': 0, 'amount': 0, 'vat': 0})
        for item in purchase_items:
            cat = item['category']
            purchase_by_category[cat]['count'] += 1
            purchase_by_category[cat]['amount'] += item['amount']
            purchase_by_category[cat]['vat'] += item['vat']
        
        return Response({
            'document_id': document.id,
            'filename': document.original_filename,
            'period': {
                'year': year,
                'quarter': quarter,
                'description': f"{year}년 {'제' + quarter + '기' if quarter else '전체'}",
            },
            'sales': {
                'total_amount': total_sales,
                'supply_value': round(total_sales * 10 / 11, 0),
                'vat': sales_vat,
                'count': len(sales_items),
                'items': sales_items[:100],  # 상위 100건
            },
            'purchases': {
                'total_amount': total_purchases,
                'supply_value': round(total_purchases * 10 / 11, 0),
                'vat': purchase_vat,
                'count': len(purchase_items),
                'by_category': dict(purchase_by_category),
                'items': purchase_items[:100],
            },
            'summary': {
                'sales_vat': sales_vat,
                'purchase_vat': purchase_vat,
                'payable_vat': payable_vat,
                'refund': max(-payable_vat, 0) if payable_vat < 0 else 0,
            }
        })

    @action(detail=True, methods=['get'])
    def vat_download(self, request, pk=None):
        """부가세 신고서 엑셀 다운로드"""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response({'error': '추출된 데이터가 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        fin_cols = detect_financial_columns(headers)
        income_idx = fin_cols.get('income')
        expense_idx = fin_cols.get('expense')
        desc_idx = fin_cols.get('description')
        date_idx = fin_cols.get('date')
        
        quarter = request.query_params.get('quarter', '')
        year = request.query_params.get('year', str(date.today().year))
        
        wb = openpyxl.Workbook()
        
        # 스타일
        title_font = Font(bold=True, size=16, color='1F2937')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        money_fmt = '#,##0'
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        
        # === 시트 1: 부가세 요약 ===
        ws = wb.active
        ws.title = '부가세 요약'
        
        period_str = f"{year}년 {'제' + quarter + '기' if quarter else '전체'}"
        ws.cell(row=1, column=1, value=f'부가가치세 신고 요약 - {period_str}').font = title_font
        ws.merge_cells('A1:D1')
        
        # 매출 세액
        ws.cell(row=3, column=1, value='■ 매출 세액').font = Font(bold=True, size=12, color='10B981')
        ws.cell(row=4, column=1, value='구분').font = Font(bold=True)
        ws.cell(row=4, column=2, value='공급가액').font = Font(bold=True)
        ws.cell(row=4, column=3, value='세액').font = Font(bold=True)
        
        total_sales = 0
        total_purchases = 0
        
        for row in rows:
            if income_idx is not None and income_idx < len(row):
                try:
                    val = row[income_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if not math.isnan(amount):
                        total_sales += amount
                except:
                    pass
            if expense_idx is not None and expense_idx < len(row):
                try:
                    val = row[expense_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if not math.isnan(amount):
                        total_purchases += amount
                except:
                    pass
        
        sales_supply = round(total_sales * 10 / 11, 0)
        sales_vat = round(total_sales / 11, 0)
        purchase_supply = round(total_purchases * 10 / 11, 0)
        purchase_vat = round(total_purchases / 11, 0)
        
        ws.cell(row=5, column=1, value='과세 매출')
        ws.cell(row=5, column=2, value=sales_supply).number_format = money_fmt
        ws.cell(row=5, column=3, value=sales_vat).number_format = money_fmt
        
        # 매입 세액
        ws.cell(row=7, column=1, value='■ 매입 세액').font = Font(bold=True, size=12, color='EF4444')
        ws.cell(row=8, column=1, value='구분').font = Font(bold=True)
        ws.cell(row=8, column=2, value='공급가액').font = Font(bold=True)
        ws.cell(row=8, column=3, value='세액').font = Font(bold=True)
        
        ws.cell(row=9, column=1, value='과세 매입')
        ws.cell(row=9, column=2, value=purchase_supply).number_format = money_fmt
        ws.cell(row=9, column=3, value=purchase_vat).number_format = money_fmt
        
        # 납부 세액
        payable = sales_vat - purchase_vat
        ws.cell(row=11, column=1, value='■ 납부(환급) 세액').font = Font(bold=True, size=12, color='2563EB')
        ws.cell(row=12, column=1, value='납부할 세액' if payable >= 0 else '환급받을 세액').font = Font(bold=True)
        ws.cell(row=12, column=2, value=abs(payable)).number_format = money_fmt
        ws.cell(row=12, column=2).font = Font(bold=True, size=14, color='2563EB')
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
        
        # === 시트 2: 매출 내역 ===
        ws2 = wb.create_sheet('매출 내역')
        sale_headers = ['날짜', '적요', '계정과목', '금액', '공급가액', '세액']
        for ci, h in enumerate(sale_headers, 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        row_num = 2
        for row in rows:
            if income_idx is not None and income_idx < len(row):
                try:
                    val = row[income_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if math.isnan(amount) or amount <= 0:
                        continue
                except:
                    continue
                
                desc = str(row[desc_idx]) if desc_idx is not None and desc_idx < len(row) else ''
                dt = str(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
                cat = classify_transaction(desc)
                
                ws2.cell(row=row_num, column=1, value=dt)
                ws2.cell(row=row_num, column=2, value=desc)
                ws2.cell(row=row_num, column=3, value=cat)
                ws2.cell(row=row_num, column=4, value=amount).number_format = money_fmt
                ws2.cell(row=row_num, column=5, value=round(amount * 10 / 11, 0)).number_format = money_fmt
                ws2.cell(row=row_num, column=6, value=round(amount / 11, 0)).number_format = money_fmt
                row_num += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2.column_dimensions[col].width = 18
        
        # === 시트 3: 매입 내역 ===
        ws3 = wb.create_sheet('매입 내역')
        for ci, h in enumerate(sale_headers, 1):
            cell = ws3.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        row_num = 2
        for row in rows:
            if expense_idx is not None and expense_idx < len(row):
                try:
                    val = row[expense_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if math.isnan(amount) or amount <= 0:
                        continue
                except:
                    continue
                
                desc = str(row[desc_idx]) if desc_idx is not None and desc_idx < len(row) else ''
                dt = str(row[date_idx]) if date_idx is not None and date_idx < len(row) else ''
                cat = classify_transaction(desc)
                
                ws3.cell(row=row_num, column=1, value=dt)
                ws3.cell(row=row_num, column=2, value=desc)
                ws3.cell(row=row_num, column=3, value=cat)
                ws3.cell(row=row_num, column=4, value=amount).number_format = money_fmt
                ws3.cell(row=row_num, column=5, value=round(amount * 10 / 11, 0)).number_format = money_fmt
                ws3.cell(row=row_num, column=6, value=round(amount / 11, 0)).number_format = money_fmt
                row_num += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws3.column_dimensions[col].width = 18
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        filename = f'부가세신고_{document.original_filename.rsplit(".", 1)[0]}_{period_str}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ========================
    # 월별 손익 리포트
    # ========================

    @action(detail=True, methods=['get'])
    def monthly_report(self, request, pk=None):
        """월별 손익 리포트 API"""
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response({'error': '추출된 데이터가 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        
        fin_cols = detect_financial_columns(headers)
        income_idx = fin_cols.get('income')
        expense_idx = fin_cols.get('expense')
        desc_idx = fin_cols.get('description')
        date_idx = fin_cols.get('date')
        
        if date_idx is None:
            return Response({'error': '날짜 열을 감지할 수 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)
        
        monthly = defaultdict(lambda: {
            'income': 0, 'expense': 0, 'net': 0, 'count': 0,
            'categories': defaultdict(lambda: {'income': 0, 'expense': 0, 'count': 0}),
        })
        
        for row in rows:
            if date_idx >= len(row) or not row[date_idx]:
                continue
            
            d = _parse_date(str(row[date_idx]))
            if not d:
                continue
            
            month_key = f"{d.year}-{d.month:02d}"
            monthly[month_key]['count'] += 1
            
            desc = str(row[desc_idx]) if desc_idx is not None and desc_idx < len(row) else ''
            category = classify_transaction(desc)
            
            inc_amount = 0
            exp_amount = 0
            
            if income_idx is not None and income_idx < len(row):
                try:
                    val = row[income_idx]
                    inc_amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if math.isnan(inc_amount):
                        inc_amount = 0
                except:
                    inc_amount = 0
            
            if expense_idx is not None and expense_idx < len(row):
                try:
                    val = row[expense_idx]
                    exp_amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if math.isnan(exp_amount):
                        exp_amount = 0
                except:
                    exp_amount = 0
            
            monthly[month_key]['income'] += inc_amount
            monthly[month_key]['expense'] += exp_amount
            monthly[month_key]['categories'][category]['income'] += inc_amount
            monthly[month_key]['categories'][category]['expense'] += exp_amount
            monthly[month_key]['categories'][category]['count'] += 1
        
        # 정렬 및 net 계산
        result = []
        for month_key in sorted(monthly.keys()):
            data = monthly[month_key]
            data['net'] = data['income'] - data['expense']
            data['month'] = month_key
            data['categories'] = dict(data['categories'])
            result.append(data)
        
        # 누적 합계
        cumulative_income = 0
        cumulative_expense = 0
        for item in result:
            cumulative_income += item['income']
            cumulative_expense += item['expense']
            item['cumulative_income'] = cumulative_income
            item['cumulative_expense'] = cumulative_expense
            item['cumulative_net'] = cumulative_income - cumulative_expense
        
        return Response({
            'document_id': document.id,
            'filename': document.original_filename,
            'months': result,
            'total': {
                'income': cumulative_income,
                'expense': cumulative_expense,
                'net': cumulative_income - cumulative_expense,
                'transaction_count': sum(m['count'] for m in result),
            }
        })

    # ========================
    # 거래처 추출
    # ========================

    @action(detail=True, methods=['post'])
    def extract_vendors(self, request, pk=None):
        """거래 내역에서 거래처를 자동 추출하여 Vendor 모델에 저장"""
        document = self.get_object()
        
        try:
            extracted = document.extracted_data
        except ExtractedData.DoesNotExist:
            return Response({'error': '추출된 데이터가 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        sd = extracted.structured_data
        headers = sd.get('headers', [])
        rows = sd.get('rows', [])
        
        fin_cols = detect_financial_columns(headers)
        desc_idx = fin_cols.get('description')
        income_idx = fin_cols.get('income')
        expense_idx = fin_cols.get('expense')
        
        if desc_idx is None:
            return Response({'error': '적요 열을 감지할 수 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 거래처별 집계
        vendor_stats = defaultdict(lambda: {
            'total_income': 0, 'total_expense': 0, 'count': 0, 'category': '미분류'
        })
        
        for row in rows:
            if desc_idx >= len(row) or not row[desc_idx]:
                continue
            
            desc = str(row[desc_idx]).strip()
            if not desc:
                continue
            
            # 거래처명 추출 (적요에서 주요 키워드 제거)
            vendor_name = _extract_vendor_name(desc)
            if not vendor_name or len(vendor_name) < 2:
                continue
            
            vendor_stats[vendor_name]['count'] += 1
            vendor_stats[vendor_name]['category'] = classify_transaction(desc)
            
            if income_idx is not None and income_idx < len(row):
                try:
                    val = row[income_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if not math.isnan(amount):
                        vendor_stats[vendor_name]['total_income'] += amount
                except:
                    pass
            
            if expense_idx is not None and expense_idx < len(row):
                try:
                    val = row[expense_idx]
                    amount = float(str(val).replace(',', '').replace('원', '').strip()) if val else 0
                    if not math.isnan(amount):
                        vendor_stats[vendor_name]['total_expense'] += amount
                except:
                    pass
        
        # Vendor 모델에 저장
        created_count = 0
        updated_count = 0
        for name, stats in vendor_stats.items():
            if stats['count'] < 1:
                continue
            
            vendor_type = 'customer' if stats['total_income'] > stats['total_expense'] else 'supplier'
            
            vendor, created = Vendor.objects.update_or_create(
                user=request.user,
                name=name,
                defaults={
                    'vendor_type': vendor_type,
                    'category': stats['category'],
                    'total_income': stats['total_income'],
                    'total_expense': stats['total_expense'],
                    'transaction_count': stats['count'],
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return Response({
            'message': f'거래처 {created_count}개 생성, {updated_count}개 갱신',
            'total_vendors': created_count + updated_count,
        })


class ExtractedDataViewSet(viewsets.ReadOnlyModelViewSet):
    """추출 데이터 뷰셋 (읽기 전용)"""
    serializer_class = ExtractedDataSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return ExtractedData.objects.filter(document__user=self.request.user)


class ReportViewSet(viewsets.ModelViewSet):
    """리포트 뷰셋"""
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['title', 'summary']
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return Report.objects.filter(document__user=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(generated_by=self.request.user)


# ========================
# 파일 병합 뷰셋
# ========================

class MergeProjectViewSet(viewsets.ModelViewSet):
    """
    파일 병합 프로젝트 뷰셋
    
    워크플로우:
    1. POST /merge-projects/ → 프로젝트 생성
    2. POST /merge-projects/{id}/upload_files/ → 파일 업로드
    3. POST /merge-projects/{id}/analyze/ → 파일 분석 (헤더 탐지, 매핑 제안)
    4. PUT /merge-projects/{id}/update_mapping/ → 매핑 규칙 확정
    5. POST /merge-projects/{id}/execute/ → 병합 실행
    6. GET /merge-projects/{id}/download/ → 결과 다운로드
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return MergeProject.objects.filter(user=self.request.user).prefetch_related('files')
    
    def get_serializer_class(self):
        if self.action == 'create':
            return MergeProjectCreateSerializer
        return MergeProjectSerializer
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_files(self, request, pk=None):
        """병합 대상 파일 업로드 (복수 파일)"""
        project = self.get_object()
        
        if project.status not in ('draft', 'ready', 'failed'):
            return Response(
                {'error': f'현재 상태({project.get_status_display()})에서는 파일을 추가할 수 없습니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        files = request.FILES.getlist('files')
        if not files:
            return Response(
                {'error': '파일을 선택해주세요.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 파일 검증
        serializer = MergeFileUploadSerializer(data={'files': files})
        serializer.is_valid(raise_exception=True)
        
        created_files = []
        for f in files:
            merge_file = MergeFile.objects.create(
                project=project,
                file=f,
                original_filename=f.name,
                file_size=f.size,
            )
            created_files.append(merge_file)
        
        # 파일 추가 후 상태를 draft으로 리셋
        if project.status != 'draft':
            project.status = 'draft'
            project.save()
        
        return Response({
            'message': f'{len(created_files)}개 파일이 업로드되었습니다.',
            'files': MergeFileSerializer(created_files, many=True).data,
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def analyze(self, request, pk=None):
        """파일 분석 시작 (1단계)"""
        project = self.get_object()
        
        if project.files.count() == 0:
            return Response(
                {'error': '분석할 파일이 없습니다. 먼저 파일을 업로드해주세요.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if project.status == 'analyzing':
            return Response(
                {'error': '이미 분석이 진행 중입니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        project.status = 'analyzing'
        project.save()
        
        _dispatch_task(analyze_merge_files, project.id)
        
        return Response({
            'message': '파일 분석이 시작되었습니다.',
            'project_id': project.id,
        })
    
    @action(detail=True, methods=['put', 'patch'])
    def update_mapping(self, request, pk=None):
        """매핑 규칙 업데이트 (2단계)"""
        project = self.get_object()
        
        serializer = MergeProjectUpdateMappingSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        
        if 'column_mapping' in data:
            project.column_mapping = data['column_mapping']
        if 'date_columns' in data:
            project.date_columns = data['date_columns']
        if 'number_columns' in data:
            project.number_columns = data['number_columns']
        if 'date_output_format' in data:
            project.date_output_format = data['date_output_format']
        
        project.status = 'ready'
        project.save()
        
        return Response({
            'message': '매핑 규칙이 업데이트되었습니다.',
            'project': MergeProjectSerializer(project, context={'request': request}).data,
        })
    
    @action(detail=True, methods=['post'])
    def apply_template(self, request, pk=None):
        """매핑 템플릿 적용"""
        project = self.get_object()
        template_id = request.data.get('template_id')
        
        if not template_id:
            return Response(
                {'error': 'template_id를 지정해주세요.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            template = ColumnMappingTemplate.objects.get(
                id=template_id,
                user=request.user
            )
        except ColumnMappingTemplate.DoesNotExist:
            # 공개 템플릿도 확인
            try:
                template = ColumnMappingTemplate.objects.get(
                    id=template_id,
                    is_public=True
                )
            except ColumnMappingTemplate.DoesNotExist:
                return Response(
                    {'error': '템플릿을 찾을 수 없습니다.'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        project.column_mapping = template.column_mapping
        project.date_columns = template.date_columns
        project.number_columns = template.number_columns
        project.date_output_format = template.date_output_format
        project.status = 'ready'
        project.save()
        
        return Response({
            'message': f'템플릿 "{template.name}"이(가) 적용되었습니다.',
            'project': MergeProjectSerializer(project, context={'request': request}).data,
        })
    
    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        """병합 실행 (3단계)"""
        project = self.get_object()
        
        if project.status not in ('ready', 'failed'):
            return Response(
                {'error': f'현재 상태({project.get_status_display()})에서는 병합을 실행할 수 없습니다. '
                          f'먼저 분석 및 매핑 규칙 설정을 완료해주세요.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        _dispatch_task(execute_merge, project.id)
        
        return Response({
            'message': '병합이 시작되었습니다. 완료 후 결과를 다운로드할 수 있습니다.',
            'project_id': project.id,
        })
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """병합 결과 파일 다운로드"""
        from django.http import FileResponse
        import os
        
        project = self.get_object()
        
        if not project.merged_file:
            return Response(
                {'error': '병합 결과 파일이 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        file_path = project.merged_file.path
        if not os.path.exists(file_path):
            return Response(
                {'error': '파일을 찾을 수 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        filename = f'{project.name}_병합결과.xlsx'
        response = FileResponse(
            open(file_path, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(detail=True, methods=['post'])
    def save_as_template(self, request, pk=None):
        """현재 프로젝트의 매핑 규칙을 템플릿으로 저장"""
        project = self.get_object()
        
        name = request.data.get('name', f'{project.name} 템플릿')
        description = request.data.get('description', '')
        
        template = ColumnMappingTemplate.objects.create(
            user=request.user,
            name=name,
            description=description,
            column_mapping=project.column_mapping,
            date_columns=project.date_columns,
            number_columns=project.number_columns,
            date_output_format=project.date_output_format,
        )
        
        return Response({
            'message': '매핑 템플릿이 저장되었습니다.',
            'template': ColumnMappingTemplateSerializer(template).data,
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['get'])
    def files(self, request, pk=None):
        """프로젝트의 파일 목록"""
        project = self.get_object()
        files = project.files.all()
        serializer = MergeFileSerializer(files, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['delete'], url_path='files/(?P<file_id>[0-9]+)')
    def remove_file(self, request, pk=None, file_id=None):
        """프로젝트에서 개별 파일 제거"""
        project = self.get_object()
        
        try:
            merge_file = project.files.get(id=file_id)
            merge_file.file.delete()  # 실제 파일 삭제
            merge_file.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except MergeFile.DoesNotExist:
            return Response(
                {'error': '파일을 찾을 수 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )


class ColumnMappingTemplateViewSet(viewsets.ModelViewSet):
    """매핑 템플릿 뷰셋"""
    serializer_class = ColumnMappingTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-updated_at']
    
    def get_queryset(self):
        from django.db.models import Q
        return ColumnMappingTemplate.objects.filter(
            Q(user=self.request.user) | Q(is_public=True)
        )


# ========================
# 거래처 관리 뷰셋
# ========================

class VendorViewSet(viewsets.ModelViewSet):
    """거래처 관리 뷰셋"""
    serializer_class = VendorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['vendor_type', 'category']
    search_fields = ['name', 'business_number', 'memo']
    ordering_fields = ['name', 'total_income', 'total_expense', 'transaction_count', 'created_at']
    ordering = ['-transaction_count']
    
    def get_queryset(self):
        return Vendor.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """거래처 요약 통계"""
        vendors = self.get_queryset()
        
        total_vendors = vendors.count()
        customers = vendors.filter(vendor_type='customer')
        suppliers = vendors.filter(vendor_type='supplier')
        
        from django.db.models import Sum
        
        return Response({
            'total_vendors': total_vendors,
            'customers': {
                'count': customers.count(),
                'total_income': customers.aggregate(s=Sum('total_income'))['s'] or 0,
            },
            'suppliers': {
                'count': suppliers.count(),
                'total_expense': suppliers.aggregate(s=Sum('total_expense'))['s'] or 0,
            },
        })


# ========================
# 세금 달력 API
# ========================

from rest_framework.decorators import api_view, permission_classes as perm_classes
from rest_framework.permissions import IsAuthenticated as IsAuth


@api_view(['GET'])
@perm_classes([IsAuth])
def tax_calendar_api(request):
    """세금 달력 API
    
    쿼리 파라미터:
    - year: 연도 (기본 현재년)
    - month: 월 (필터, 미지정시 전체)
    """
    year = int(request.query_params.get('year', date.today().year))
    month = request.query_params.get('month')
    
    events = []
    today = date.today()
    
    for item in TAX_CALENDAR:
        if month and item['month'] != int(month):
            continue
        
        try:
            event_date = date(year, item['month'], item['day'])
        except ValueError:
            continue
        
        days_until = (event_date - today).days
        
        events.append({
            'date': event_date.isoformat(),
            'month': item['month'],
            'day': item['day'],
            'title': item['title'],
            'description': item['desc'],
            'type': item['type'],
            'days_until': days_until,
            'status': 'overdue' if days_until < 0 else ('upcoming' if days_until <= 7 else 'future'),
        })
    
    # 가까운 일정 순서
    events.sort(key=lambda x: x['date'])
    
    # 다음 다가오는 일정
    upcoming = [e for e in events if e['days_until'] >= 0]
    
    return Response({
        'year': year,
        'events': events,
        'next_event': upcoming[0] if upcoming else None,
        'upcoming_count': len([e for e in upcoming if e['days_until'] <= 30]),
    })
