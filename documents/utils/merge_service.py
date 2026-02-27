"""
다중 엑셀 파일 병합 서비스

헤더 탐지, 열 매핑, 포맷 정규화를 결합하여
구조가 비슷한 여러 엑셀 파일을 하나로 병합합니다.
"""
import os
import json
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class MergeService:
    """다중 엑셀 파일 병합 서비스"""
    
    def __init__(self):
        from .normalizers import DateNormalizer, NumberNormalizer
        from .header_detector import HeaderDetector
        from .column_mapper import ColumnMapper
        
        self.date_normalizer = DateNormalizer()
        self.number_normalizer = NumberNormalizer()
        self.header_detector = HeaderDetector()
        self.column_mapper = ColumnMapper()
    
    def analyze_files(self, file_paths: List[str]) -> dict:
        """
        여러 파일을 분석하여 구조 정보와 매핑 제안을 반환
        (1단계: 샘플 분석)
        
        Args:
            file_paths: 분석할 엑셀 파일 경로 목록
            
        Returns:
            분석 결과 딕셔너리
        """
        import openpyxl
        
        analysis = {
            'files': [],
            'suggested_mappings': {},
            'all_headers': [],
            'common_structure': {},
        }
        
        all_headers_list = []
        
        for file_path in file_paths:
            try:
                file_info = self._analyze_single_file(file_path)
                analysis['files'].append(file_info)
                all_headers_list.append(file_info['headers'])
            except Exception as e:
                logger.error(f"파일 분석 오류 ({file_path}): {e}")
                analysis['files'].append({
                    'file_path': file_path,
                    'filename': os.path.basename(file_path),
                    'error': str(e),
                })
        
        # 매핑 제안 생성
        if all_headers_list:
            analysis['suggested_mappings'] = self.column_mapper.suggest_mappings(all_headers_list)
            
            # 모든 고유 헤더 수집
            all_unique = set()
            for headers in all_headers_list:
                all_unique.update(headers)
            analysis['all_headers'] = sorted(list(all_unique))
        
        return analysis
    
    def _analyze_single_file(self, file_path: str) -> dict:
        """단일 파일 분석"""
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        
        # 전체 데이터 읽기
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        
        # 헤더 탐지
        headers, data_rows, meta_info = self.header_detector.extract_data_with_header(rows)
        
        # 각 열의 데이터 타입 분석
        column_types = self._analyze_column_types(headers, data_rows)
        
        # 샘플 데이터 (처음 5행)
        sample_data = data_rows[:5]
        
        wb.close()
        
        return {
            'file_path': file_path,
            'filename': os.path.basename(file_path),
            'headers': headers,
            'header_row_index': meta_info['header_row_index'],
            'total_rows': len(data_rows),
            'column_types': column_types,
            'sample_data': self._serialize_data(sample_data),
            'meta_info': meta_info,
        }
    
    def _analyze_column_types(self, headers: List[str], data_rows: List[List[Any]]) -> Dict[str, str]:
        """각 열의 데이터 타입을 분석"""
        column_types = {}
        
        for col_idx, header in enumerate(headers):
            values = []
            for row in data_rows[:50]:  # 최대 50행 분석
                if col_idx < len(row) and row[col_idx] is not None:
                    values.append(row[col_idx])
            
            if not values:
                column_types[header] = 'empty'
                continue
            
            # 타입 비율 계산
            type_counts = {'number': 0, 'date': 0, 'text': 0}
            
            for v in values:
                if isinstance(v, (int, float)):
                    type_counts['number'] += 1
                elif isinstance(v, datetime):
                    type_counts['date'] += 1
                elif isinstance(v, str):
                    # 숫자로 파싱 가능 여부
                    normalized_num = self.number_normalizer.normalize(v)
                    if normalized_num is not None:
                        type_counts['number'] += 1
                    else:
                        # 날짜로 파싱 가능 여부
                        normalized_date = self.date_normalizer.normalize(v)
                        if normalized_date is not None:
                            type_counts['date'] += 1
                        else:
                            type_counts['text'] += 1
                else:
                    type_counts['text'] += 1
            
            # 최다 타입 선택
            dominant_type = max(type_counts, key=type_counts.get)
            column_types[header] = dominant_type
        
        return column_types
    
    def merge_files(self, 
                    file_paths: List[str],
                    column_mapping: Optional[Dict[str, str]] = None,
                    date_columns: Optional[List[str]] = None,
                    number_columns: Optional[List[str]] = None,
                    output_path: Optional[str] = None,
                    add_source_column: bool = True,
                    sort_by: Optional[str] = None,
                    auto_detect_types: bool = True) -> dict:
        """
        여러 엑셀 파일을 하나로 병합
        (2단계: 실행)
        
        Args:
            file_paths: 병합할 파일 경로 목록
            column_mapping: 열 이름 매핑 {원본명: 표준명}
            date_columns: 날짜 정규화 대상 열 이름 목록 (비어있으면 자동 감지)
            number_columns: 숫자 정규화 대상 열 이름 목록 (비어있으면 자동 감지)
            output_path: 결과 파일 경로 (None이면 자동 생성)
            add_source_column: 소스 파일명 열 추가 여부
            sort_by: 정렬 기준 열 이름 (None이면 날짜 열 자동 감지)
            auto_detect_types: 날짜/숫자 열 자동 감지 여부
            
        Returns:
            병합 결과 정보
        """
        import openpyxl
        
        if column_mapping is None:
            column_mapping = {}
        if date_columns is None:
            date_columns = []
        if number_columns is None:
            number_columns = []
        
        # 자동 타입 감지: 분석 결과에서 date/number 열을 자동으로 추가
        auto_detected = {'date_columns': [], 'number_columns': [], 'sort_column': None}
        if auto_detect_types and not date_columns:
            # 먼저 분석하여 타입 정보 수집
            analysis = self.analyze_files(file_paths)
            for file_info in analysis.get('files', []):
                col_types = file_info.get('column_types', {})
                for col_name, col_type in col_types.items():
                    if col_type == 'date' and col_name not in date_columns:
                        if col_name not in auto_detected['date_columns']:
                            auto_detected['date_columns'].append(col_name)
                    elif col_type == 'number' and col_name not in number_columns:
                        if col_name not in auto_detected['number_columns']:
                            auto_detected['number_columns'].append(col_name)
            
            date_columns = list(set(date_columns + auto_detected['date_columns']))
            number_columns = list(set(number_columns + auto_detected['number_columns']))
            
            # 날짜 열 중 정렬 기준 자동 결정
            if not sort_by and date_columns:
                date_hints = ['날짜', '거래일', '일자', '거래일자', 'date', '시간', '이용일']
                for hint in date_hints:
                    for dc in date_columns:
                        if hint in dc.lower():
                            auto_detected['sort_column'] = dc
                            break
                    if auto_detected['sort_column']:
                        break
                if not auto_detected['sort_column']:
                    auto_detected['sort_column'] = date_columns[0]
                sort_by = auto_detected['sort_column']
        
        logger.info(f"병합 설정: date_columns={date_columns}, number_columns={number_columns}, sort_by={sort_by}")
        
        all_data = []
        merge_log = []
        errors = []
        
        for file_path in file_paths:
            try:
                result = self._process_single_file(
                    file_path, column_mapping, date_columns, number_columns
                )
                
                filename = os.path.basename(file_path)
                
                for row_data in result['data']:
                    if add_source_column:
                        row_data['__source_file__'] = filename
                    all_data.append(row_data)
                
                merge_log.append({
                    'file': filename,
                    'rows_processed': result['rows_processed'],
                    'header_row': result['header_row_index'],
                    'original_headers': result['original_headers'],
                    'mapped_headers': result['mapped_headers'],
                    'status': 'success',
                })
                
            except Exception as e:
                logger.error(f"파일 병합 오류 ({file_path}): {e}")
                errors.append({
                    'file': os.path.basename(file_path),
                    'error': str(e),
                })
                merge_log.append({
                    'file': os.path.basename(file_path),
                    'status': 'error',
                    'error': str(e),
                })
        
        if not all_data:
            return {
                'success': False,
                'error': '병합할 데이터가 없습니다.',
                'merge_log': merge_log,
                'errors': errors,
            }
        
        # 통합 헤더 생성
        unified_headers = self._build_unified_headers(all_data, add_source_column)
        
        # ★ 정렬: sort_by 열 기준으로 전체 데이터 정렬
        if sort_by and sort_by in unified_headers:
            logger.info(f"병합 결과를 '{sort_by}' 기준으로 정렬합니다.")
            def _sort_key(row):
                val = row.get(sort_by, '')
                if val is None:
                    return ''
                return str(val)
            all_data.sort(key=_sort_key)
        
        # ★ 중복 탐지: 같은 날짜+금액+적요 조합이면 의심 중복
        duplicates = self._detect_duplicates(all_data, unified_headers)
        
        # 출력 파일 생성
        if output_path is None:
            output_dir = os.path.dirname(file_paths[0]) if file_paths else '.'
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(output_dir, f'merged_{timestamp}.xlsx')
        
        self._write_output(all_data, unified_headers, output_path)
        
        return {
            'success': True,
            'output_path': output_path,
            'total_rows': len(all_data),
            'total_files': len(file_paths),
            'files_succeeded': len(file_paths) - len(errors),
            'files_failed': len(errors),
            'unified_headers': unified_headers,
            'merge_log': merge_log,
            'errors': errors,
            'auto_detected': auto_detected,
            'sorted_by': sort_by,
            'duplicates': duplicates,
        }
    
    def _process_single_file(self, file_path: str, 
                              column_mapping: Dict[str, str],
                              date_columns: List[str],
                              number_columns: List[str]) -> dict:
        """단일 파일 처리하여 정규화된 데이터 반환"""
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        
        wb.close()
        
        # 헤더 탐지
        headers, data_rows, meta_info = self.header_detector.extract_data_with_header(rows)
        
        # 열 이름 매핑
        mapped_headers = []
        for h in headers:
            if h in column_mapping:
                mapped_headers.append(column_mapping[h])
            else:
                standard_name, confidence = self.column_mapper.map_column(h)
                mapped_headers.append(standard_name if confidence > 0.5 else h)
        
        # 데이터 정규화
        processed_data = []
        for row in data_rows:
            row_dict = {}
            for col_idx, header in enumerate(mapped_headers):
                value = row[col_idx] if col_idx < len(row) else None
                
                # 날짜 정규화
                if header in date_columns:
                    value = self.date_normalizer.normalize(value) or value
                
                # 숫자 정규화
                if header in number_columns:
                    normalized = self.number_normalizer.normalize(value)
                    if normalized is not None:
                        value = normalized
                
                # JSON 직렬화 가능한 값으로 변환
                value = self._serialize_value(value)
                row_dict[header] = value
            
            processed_data.append(row_dict)
        
        return {
            'data': processed_data,
            'rows_processed': len(processed_data),
            'header_row_index': meta_info['header_row_index'],
            'original_headers': headers,
            'mapped_headers': mapped_headers,
        }
    
    def _detect_duplicates(self, all_data: List[Dict], headers: List[str]) -> dict:
        """중복 행 탐지 — 같은 값 조합이 여러 소스 파일에서 나타나면 의심 중복"""
        seen = {}  # signature -> [indices]
        duplicates = []
        
        # 서명 생성 열: __source_file__ 제외한 모든 값
        for idx, row in enumerate(all_data):
            sig_parts = []
            for h in headers:
                if h == '__source_file__':
                    continue
                val = row.get(h, '')
                sig_parts.append(str(val) if val is not None else '')
            signature = '|'.join(sig_parts)
            
            if signature in seen:
                seen[signature].append(idx)
            else:
                seen[signature] = [idx]
        
        for sig, indices in seen.items():
            if len(indices) > 1:
                # 서로 다른 소스 파일에서 왔는지 확인
                sources = set()
                for i in indices:
                    src = all_data[i].get('__source_file__', f'row_{i}')
                    sources.add(src)
                
                if len(sources) > 1:  # 다른 파일에서 같은 데이터 = 실제 중복 가능성
                    duplicates.append({
                        'rows': indices,
                        'sources': list(sources),
                        'count': len(indices),
                        'sample': sig[:200],
                    })
        
        return {
            'total_suspected': len(duplicates),
            'details': duplicates[:50],  # 최대 50건
        }
    
    def _build_unified_headers(self, all_data: List[Dict], add_source: bool) -> List[str]:
        """모든 데이터에서 통합 헤더 생성"""
        header_set = set()
        header_order = []
        
        for row in all_data:
            for key in row.keys():
                if key not in header_set:
                    header_set.add(key)
                    header_order.append(key)
        
        # 소스 파일 열을 맨 앞으로
        if add_source and '__source_file__' in header_order:
            header_order.remove('__source_file__')
            header_order.insert(0, '__source_file__')
        
        return header_order
    
    def _write_output(self, data: List[Dict], headers: List[str], output_path: str):
        """병합된 데이터를 엑셀 파일로 출력"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '병합 결과'
        
        # 헤더 스타일
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # 헤더 이름 한글화
        display_headers = []
        for h in headers:
            if h == '__source_file__':
                display_headers.append('원본 파일')
            else:
                display_headers.append(h)
        
        # 헤더 행 작성
        for col_idx, header in enumerate(display_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 데이터 행 작성
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # 열 너비 자동 조정
        for col_idx, header in enumerate(display_headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(data) + 2, 52)):  # 최대 50행 체크
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)
        
        # 필터 설정
        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}1'
        
        # 행 고정
        ws.freeze_panes = 'A2'
        
        # 저장
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        
        logger.info(f"병합 파일 저장: {output_path} ({len(data)}행)")
    
    def _serialize_value(self, value: Any) -> Any:
        """값을 JSON 직렬화 가능한 형태로 변환"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, (int, float, str, bool)):
            return value
        return str(value)
    
    def _serialize_data(self, data: List[List[Any]]) -> List[List[Any]]:
        """2D 리스트 직렬화"""
        return [
            [self._serialize_value(cell) for cell in row]
            for row in data
        ]
