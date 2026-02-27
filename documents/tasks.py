from celery import shared_task
from django.utils import timezone
from .models import Document, ExtractedData, Report, MergeProject, MergeFile
import logging
import os

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def process_document(self, document_id):
    """문서 처리 태스크"""
    try:
        document = Document.objects.get(id=document_id)
        document.status = 'processing'
        document.save()
        
        # 파일 유형에 따라 처리
        if document.file_type == 'excel':
            result = process_excel(document)
        elif document.file_type == 'image':
            result = process_image(document)
        elif document.file_type == 'pdf':
            result = process_pdf(document)
        else:
            raise ValueError(f"지원하지 않는 파일 유형: {document.file_type}")
        
        # 추출된 데이터 저장
        ExtractedData.objects.update_or_create(
            document=document,
            defaults=result
        )
        
        # 문서 상태 업데이트
        document.status = 'completed'
        document.processed_at = timezone.now()
        document.save()
        
        # 리포트 생성
        generate_report.delay(document_id)
        
        logger.info(f"문서 처리 완료: {document.original_filename}")
        return {'status': 'success', 'document_id': document_id}
        
    except Document.DoesNotExist:
        logger.error(f"문서를 찾을 수 없습니다: {document_id}")
        raise
    except Exception as e:
        logger.error(f"문서 처리 오류: {str(e)}")
        document.status = 'failed'
        document.error_message = str(e)
        document.save()
        
        # 재시도
        raise self.retry(exc=e, countdown=60)


def process_excel(document):
    """엑셀 파일 처리 — 전처리 파이프라인 적용
    
    1. 전체 시트 데이터 읽기
    2. HeaderDetector로 실제 헤더 행 탐지 (은행/카드 엑셀 대응)
    3. ColumnMapper로 표준 열 이름 매핑
    4. DateNormalizer로 날짜 통일 (YYYY-MM-DD)
    5. NumberNormalizer로 금액 통일 (float)
    6. 재무 열 자동 감지 → 날짜 기준 정렬
    7. 계정과목 자동 분류
    """
    try:
        import openpyxl
        import math
        from .utils.header_detector import HeaderDetector
        from .utils.column_mapper import ColumnMapper
        from .utils.normalizers import DateNormalizer, NumberNormalizer
        
        wb = openpyxl.load_workbook(document.file.path)
        sheet = wb.active
        
        # 1) 전체 raw 데이터 읽기
        raw_rows = []
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(list(row))
        
        if not raw_rows:
            return {
                'extracted_text': '',
                'structured_data': {'headers': [], 'rows': [], 'sheet_name': sheet.title},
                'total_rows': 0,
                'metadata': {'sheet_count': len(wb.sheetnames), 'sheet_names': wb.sheetnames},
            }
        
        # 2) 헤더 탐지 — 첫 행이 헤더가 아닐 수 있음 (은행 엑셀 등)
        detector = HeaderDetector()
        headers, data_rows, meta_info = detector.extract_data_with_header(raw_rows)
        
        # 3) 표준 열 이름 매핑
        mapper = ColumnMapper()
        mapped_headers = []
        mapping_log = {}
        for h in headers:
            standard_name, confidence = mapper.map_column(h)
            if confidence > 0.5:
                mapped_headers.append(standard_name)
                if standard_name != h:
                    mapping_log[h] = {'mapped_to': standard_name, 'confidence': round(confidence, 2)}
            else:
                mapped_headers.append(h)
        
        # 4) 재무 열 자동 감지
        FINANCIAL_PATTERNS = {
            'income': ['입금', '입금액', '수입', '매출', '대변', 'credit', 'income', '받은금액'],
            'expense': ['출금', '출금액', '지출', '지급', '차변', 'debit', 'expense', '보낸금액', '사용금액', '결제금액', '이용금액'],
            'balance': ['잔액', '잔고', '누적잔액', 'balance', '계좌잔액'],
            'date': ['날짜', '거래일', '일자', '거래일자', '사용일', '이용일', '승인일', '결제일', 'date', '거래일시'],
            'description': ['적요', '내용', '거래내용', '비고', '메모', '상세', '거래처', '이용내역', '사용처', '가맹점', '거래적요'],
        }
        
        fin_cols = {}
        normalized_headers = [str(h).strip().lower() for h in mapped_headers]
        for col_type, patterns in FINANCIAL_PATTERNS.items():
            for i, header in enumerate(normalized_headers):
                for pattern in patterns:
                    if pattern.lower() in header:
                        fin_cols[col_type] = i
                        break
                if col_type in fin_cols:
                    break
        
        # 5) DateNormalizer + NumberNormalizer 적용
        date_normalizer = DateNormalizer(output_format='%Y-%m-%d')
        number_normalizer = NumberNormalizer()
        
        date_col_indices = set()
        number_col_indices = set()
        
        # 날짜/숫자 열 자동 감지 (첫 50행 샘플링)
        for col_idx in range(len(mapped_headers)):
            date_count = 0
            number_count = 0
            total_non_empty = 0
            
            for row in data_rows[:50]:
                if col_idx >= len(row) or row[col_idx] is None:
                    continue
                val = row[col_idx]
                total_non_empty += 1
                
                # datetime 객체 자체도 날짜
                from datetime import datetime as dt_cls, date as d_cls
                if isinstance(val, (dt_cls, d_cls)):
                    date_count += 1
                elif isinstance(val, str):
                    if date_normalizer.normalize(val) is not None:
                        date_count += 1
                    elif number_normalizer.normalize(val) is not None:
                        number_count += 1
                elif isinstance(val, (int, float)):
                    number_count += 1
            
            if total_non_empty > 0:
                if date_count / total_non_empty > 0.5:
                    date_col_indices.add(col_idx)
                elif number_count / total_non_empty > 0.5:
                    number_col_indices.add(col_idx)
        
        # 재무 열로 감지된 것도 포함
        if 'date' in fin_cols:
            date_col_indices.add(fin_cols['date'])
        for key in ['income', 'expense', 'balance']:
            if key in fin_cols:
                number_col_indices.add(fin_cols[key])
        
        # 6) 정규화 적용
        normalized_rows = []
        for row in data_rows:
            new_row = list(row)
            # 행 길이를 헤더 수에 맞춤
            while len(new_row) < len(mapped_headers):
                new_row.append(None)
            
            for col_idx in date_col_indices:
                if col_idx < len(new_row) and new_row[col_idx] is not None:
                    normalized = date_normalizer.normalize(new_row[col_idx])
                    if normalized is not None:
                        new_row[col_idx] = normalized
            
            for col_idx in number_col_indices:
                if col_idx < len(new_row) and new_row[col_idx] is not None:
                    normalized = number_normalizer.normalize(new_row[col_idx])
                    if normalized is not None:
                        new_row[col_idx] = normalized
            
            # None 이외의 데이터가 하나라도 있는 행만 포함 (빈 행 제거)
            if any(cell is not None and str(cell).strip() != '' for cell in new_row[:len(mapped_headers)]):
                normalized_rows.append(new_row)
        
        # 7) 날짜 기준 정렬
        date_sort_idx = fin_cols.get('date')
        if date_sort_idx is not None:
            def sort_key(row):
                if date_sort_idx < len(row) and row[date_sort_idx]:
                    return str(row[date_sort_idx])
                return ''
            normalized_rows.sort(key=sort_key)
        
        # 8) 잔액 검증 (있으면)
        balance_check = None
        if all(k in fin_cols for k in ['income', 'expense', 'balance']):
            balance_idx = fin_cols['balance']
            income_idx = fin_cols['income']
            expense_idx = fin_cols['expense']
            mismatches = []
            
            for i in range(1, len(normalized_rows)):
                prev_row = normalized_rows[i - 1]
                curr_row = normalized_rows[i]
                
                try:
                    prev_bal = float(prev_row[balance_idx]) if prev_row[balance_idx] else None
                    curr_bal = float(curr_row[balance_idx]) if curr_row[balance_idx] else None
                    curr_inc = float(curr_row[income_idx]) if curr_row[income_idx] else 0
                    curr_exp = float(curr_row[expense_idx]) if curr_row[expense_idx] else 0
                    
                    if prev_bal is not None and curr_bal is not None:
                        expected = prev_bal + curr_inc - curr_exp
                        if abs(expected - curr_bal) > 0.01:
                            mismatches.append({
                                'row': i + 1,
                                'expected': expected,
                                'actual': curr_bal,
                                'diff': round(curr_bal - expected, 2),
                            })
                except (ValueError, TypeError, IndexError):
                    pass
            
            balance_check = {
                'verified': len(mismatches) == 0,
                'total_rows_checked': len(normalized_rows) - 1,
                'mismatches': mismatches[:20],  # 최대 20건
            }
        
        # 9) 계정과목 자동 분류 (적요 열이 있으면)
        ACCOUNT_KEYWORDS = {
            '매출': ['매출', '판매', '수입', '용역'],
            '이자수익': ['이자수입', '예금이자', '이자수익'],
            '급여': ['급여', '월급', '상여', '보너스', '인건비'],
            '복리후생비': ['복리후생', '식대', '직원식대', '경조사'],
            '소모품비': ['소모품', '사무용품', '비품'],
            '접대비': ['접대', '식비접대', '거래처 식비', '거래처 선물'],
            '여비교통비': ['교통', '출장', '택시', '주유', 'KTX'],
            '통신비': ['통신', '전화', '인터넷', '핸드폰'],
            '수도광열비': ['수도', '전기', '가스', '전기요금'],
            '임차료': ['임대료', '월세', '임차', '사무실 임대'],
            '보험료': ['보험', '산재', '고용보험', '건강보험', '국민연금'],
            '세금과공과': ['세금', '부가세', '법인세', '국세', '지방세'],
            '광고선전비': ['광고', '홍보', '마케팅'],
            '수수료': ['수수료', '카드수수료', '이체수수료'],
            '이자비용': ['이자', '대출이자'],
            '외주비': ['외주', '용역비', '하도급'],
        }
        
        auto_classifications = {}
        desc_idx = fin_cols.get('description')
        if desc_idx is not None:
            for row_idx, row in enumerate(normalized_rows):
                if desc_idx < len(row) and row[desc_idx]:
                    desc = str(row[desc_idx]).strip()
                    category = '미분류'
                    for cat, keywords in ACCOUNT_KEYWORDS.items():
                        if any(kw in desc for kw in keywords):
                            category = cat
                            break
                    if category != '미분류':
                        auto_classifications[str(row_idx)] = category
        
        # 구조화된 데이터 조립
        structured_data = {
            'headers': mapped_headers,
            'original_headers': headers,
            'rows': normalized_rows,
            'sheet_name': sheet.title,
        }
        
        # 메타데이터
        preprocessing_info = {
            'header_row_detected': meta_info['header_row_index'],
            'meta_rows': meta_info.get('meta_rows', []),
            'column_mapping': mapping_log,
            'date_columns': [mapped_headers[i] for i in date_col_indices if i < len(mapped_headers)],
            'number_columns': [mapped_headers[i] for i in number_col_indices if i < len(mapped_headers)],
            'financial_columns': {k: mapped_headers[v] if v < len(mapped_headers) else None for k, v in fin_cols.items()},
            'sorted_by': mapped_headers[date_sort_idx] if date_sort_idx is not None and date_sort_idx < len(mapped_headers) else None,
            'balance_check': balance_check,
            'auto_classifications': auto_classifications,
            'rows_before_cleanup': len(data_rows),
            'rows_after_cleanup': len(normalized_rows),
            'preprocessing_version': 2,
        }
        
        wb.close()
        
        return {
            'extracted_text': '',  # raw text 불필요 — structured_data가 정규화된 데이터
            'structured_data': structured_data,
            'total_rows': len(normalized_rows),
            'metadata': {
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
                'preprocessing': preprocessing_info,
            }
        }
    except Exception as e:
        logger.error(f"엑셀 처리 오류: {str(e)}")
        raise


def process_image(document):
    """이미지 파일 처리"""
    try:
        from PIL import Image
        
        img = Image.open(document.file.path)
        
        # 이미지 정보 추출
        metadata = {
            'format': img.format,
            'mode': img.mode,
            'size': img.size,
            'width': img.width,
            'height': img.height,
        }
        
        # OCR이 필요한 경우 여기에 구현
        # pytesseract 등을 사용하여 텍스트 추출
        
        return {
            'extracted_text': '',  # OCR 결과가 들어갈 자리
            'structured_data': {},
            'metadata': metadata,
        }
    except Exception as e:
        logger.error(f"이미지 처리 오류: {str(e)}")
        raise


def process_pdf(document):
    """PDF 파일 처리"""
    try:
        from PyPDF2 import PdfReader
        
        reader = PdfReader(document.file.path)
        
        # 텍스트 추출
        text = ''
        for page in reader.pages:
            text += page.extract_text() + '\n'
        
        return {
            'extracted_text': text,
            'structured_data': {},
            'total_pages': len(reader.pages),
            'metadata': {
                'page_count': len(reader.pages),
                'metadata': reader.metadata if reader.metadata else {},
            }
        }
    except Exception as e:
        logger.error(f"PDF 처리 오류: {str(e)}")
        raise


@shared_task
def generate_report(document_id):
    """리포트 생성 태스크"""
    try:
        document = Document.objects.get(id=document_id)
        extracted_data = ExtractedData.objects.get(document=document)
        
        # 리포트 생성 로직
        title = f"{document.original_filename} 분석 리포트"
        
        # 요약 생성
        summary = generate_summary(extracted_data)
        
        # 리포트 내용 생성
        content = {
            'file_info': {
                'filename': document.original_filename,
                'file_type': document.get_file_type_display(),
                'file_size': document.file_size,
                'processed_at': document.processed_at.isoformat() if document.processed_at else None,
            },
            'extracted_data': {
                'total_pages': extracted_data.total_pages,
                'total_rows': extracted_data.total_rows,
                'text_length': len(extracted_data.extracted_text),
            },
            'structured_data': extracted_data.structured_data,
        }
        
        # 리포트 저장
        Report.objects.create(
            document=document,
            title=title,
            summary=summary,
            content=content,
            generated_by=document.user,
        )
        
        logger.info(f"리포트 생성 완료: {document.original_filename}")
        
    except Exception as e:
        logger.error(f"리포트 생성 오류: {str(e)}")
        raise


def generate_summary(extracted_data):
    """요약 생성"""
    summary_parts = []
    
    if extracted_data.total_pages > 0:
        summary_parts.append(f"총 {extracted_data.total_pages}페이지")
    
    if extracted_data.total_rows > 0:
        summary_parts.append(f"총 {extracted_data.total_rows}행의 데이터")
    
    text_length = len(extracted_data.extracted_text)
    if text_length > 0:
        summary_parts.append(f"{text_length:,}자의 텍스트 추출됨")
    
    return ", ".join(summary_parts) if summary_parts else "데이터 추출 완료"


# ========================
# 파일 병합 관련 태스크
# ========================

@shared_task(bind=True, max_retries=2)
def analyze_merge_files(self, project_id):
    """병합 프로젝트의 파일들을 분석하는 태스크 (1단계)"""
    try:
        project = MergeProject.objects.get(id=project_id)
        project.status = 'analyzing'
        project.save()
        
        from .utils.merge_service import MergeService
        service = MergeService()
        
        merge_files = project.files.all()
        file_paths = [mf.file.path for mf in merge_files]
        
        # 전체 분석
        analysis = service.analyze_files(file_paths)
        
        # 개별 파일 분석 결과 저장
        for file_info in analysis.get('files', []):
            try:
                info_filename = file_info.get('filename', '')
                # 파일 경로 기반으로 매칭 (업로드 시 파일명이 변경될 수 있음)
                merge_file = None
                for mf in merge_files:
                    basename = os.path.basename(mf.file.path)
                    if basename == info_filename or mf.original_filename == info_filename:
                        merge_file = mf
                        break
                
                if merge_file is None:
                    logger.warning(f"병합 파일을 찾을 수 없음: {info_filename}")
                    continue
                
                if 'error' not in file_info:
                    merge_file.detected_headers = file_info.get('headers', [])
                    merge_file.header_row_index = file_info.get('header_row_index', 0)
                    merge_file.total_rows = file_info.get('total_rows', 0)
                    merge_file.column_types = file_info.get('column_types', {})
                    merge_file.sample_data = file_info.get('sample_data', [])
                    merge_file.is_analyzed = True
                else:
                    merge_file.error_message = file_info['error']
                merge_file.save()
            except Exception as exc:
                logger.warning(f"병합 파일 처리 오류: {file_info.get('filename')}: {exc}")
        
        # 프로젝트에 분석 결과 저장
        project.analysis_result = {
            'suggested_mappings': analysis.get('suggested_mappings', {}),
            'all_headers': analysis.get('all_headers', []),
            'analyzed_at': timezone.now().isoformat(),
            'files_analyzed': len([f for f in analysis.get('files', []) if 'error' not in f]),
            'files_failed': len([f for f in analysis.get('files', []) if 'error' in f]),
        }
        project.status = 'ready'
        project.save()
        
        logger.info(f"병합 프로젝트 분석 완료: {project.name} ({len(file_paths)}개 파일)")
        return {'status': 'success', 'project_id': project_id}
        
    except MergeProject.DoesNotExist:
        logger.error(f"병합 프로젝트를 찾을 수 없음: {project_id}")
        raise
    except Exception as e:
        logger.error(f"병합 분석 오류: {str(e)}")
        try:
            project.status = 'failed'
            project.error_message = str(e)
            project.save()
        except Exception:
            pass
        raise self.retry(exc=e, countdown=30)


@shared_task(bind=True, max_retries=2)
def execute_merge(self, project_id):
    """병합 실행 태스크 (2단계)"""
    try:
        project = MergeProject.objects.get(id=project_id)
        
        if project.status not in ('ready', 'failed'):
            logger.warning(f"병합 프로젝트 상태가 올바르지 않음: {project.status}")
            return {'status': 'skipped', 'reason': f'status={project.status}'}
        
        project.status = 'merging'
        project.error_message = ''
        project.save()
        
        from .utils.merge_service import MergeService
        from .utils.normalizers import DateNormalizer
        service = MergeService()
        
        # 날짜 포맷 설정
        if project.date_output_format:
            service.date_normalizer = DateNormalizer(output_format=project.date_output_format)
        
        # 커스텀 매핑이 있으면 적용
        if project.column_mapping:
            for original, standard in project.column_mapping.items():
                service.column_mapper.add_mapping(standard, [original])
        
        merge_files = project.files.all()
        file_paths = [mf.file.path for mf in merge_files]
        
        # 병합 실행
        from django.conf import settings as django_settings
        output_dir = os.path.join(
            str(django_settings.MEDIA_ROOT), 'merged',
            timezone.now().strftime('%Y/%m/%d')
        )
        os.makedirs(output_dir, exist_ok=True)
        
        output_filename = f'merged_{project.id}_{timezone.now().strftime("%H%M%S")}.xlsx'
        output_path = os.path.join(output_dir, output_filename)
        
        result = service.merge_files(
            file_paths=file_paths,
            column_mapping=project.column_mapping or {},
            date_columns=project.date_columns or [],
            number_columns=project.number_columns or [],
            output_path=output_path,
            add_source_column=True,
            auto_detect_types=True,  # ★ 날짜/숫자 자동 감지 활성화
        )
        
        if result['success']:
            # 병합 파일 경로를 Django FileField 상대 경로로 변환
            relative_path = os.path.relpath(output_path, str(django_settings.MEDIA_ROOT))
            project.merged_file = relative_path
            project.merge_log = result
            project.status = 'completed'
            project.completed_at = timezone.now()
            
            # 개별 파일 상태 업데이트
            for log_entry in result.get('merge_log', []):
                log_file = log_entry.get('file', '')
                mf = None
                for candidate in merge_files:
                    basename = os.path.basename(candidate.file.path)
                    if basename == log_file or candidate.original_filename == log_file:
                        mf = candidate
                        break
                if mf:
                    mf.is_processed = log_entry.get('status') == 'success'
                    if log_entry.get('error'):
                        mf.error_message = log_entry['error']
                    mf.save()
        else:
            project.status = 'failed'
            project.error_message = result.get('error', '병합 실패')
            project.merge_log = result
        
        project.save()
        
        logger.info(f"병합 실행 완료: {project.name} → {output_path}")
        return {'status': 'success', 'project_id': project_id, 'output': output_path}
        
    except MergeProject.DoesNotExist:
        logger.error(f"병합 프로젝트를 찾을 수 없음: {project_id}")
        raise
    except Exception as e:
        logger.error(f"병합 실행 오류: {str(e)}")
        try:
            project.status = 'failed'
            project.error_message = str(e)
            project.save()
        except Exception:
            pass
        raise self.retry(exc=e, countdown=30)
